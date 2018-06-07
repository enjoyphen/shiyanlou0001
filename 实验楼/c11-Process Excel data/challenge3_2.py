#!/usr/bin/env python3
# -*- coding:utf-8 -*-

import os
from collections import defaultdict
import openpyxl

def rr():
    fi_name = 'courses.xlsx'
    fi_path = os.path.join(os.path.abspath(os.path.dirname(__name__)),fi_name)

# create 'courses.xlsx' workbook instance
    wb0 = openpyxl.load_workbook(fi_path)

# return a default value of dict key
    def dict_default():
        return {'c_time':None, 'stu':0, 'time':0}

# create a dict used to save data
    dict_combine = defaultdict(dict_default)

# get data from 'students' worksheet
    for row in wb0['students'].iter_rows(min_row=2):
        dict_combine[row[1].value]['c_time'] = row[0].value
        dict_combine[row[1].value]['stu'] = row[2].value

# get data from 'time' worksheet
    for row in wb0['students'].iter_rows(min_row=2):
        dict_combine[row[1].value]['time'] = row[2].value

# create a dict to save title of workbook
    dict_wbtitle = {}

#create a 'combine' worksheet to save data
    ws_combine = wb0.create_sheet('combine')
    fi_row = ['a','b','c','d']
    ws_combine.append(fi_row)

#save to the worksheet
    for k,v in dict_combine.items():
        row = [v['c_time'], k, v['stu'], v['time']]
        ws_combine.append(row)
        title = str(v['c_time'].year)
        try:
            dict_wbtitle[title][title].append(row)
        except KeyError:
            dict_wbtitle[title] = openpyxl.Workbook()
            dict_wbtitle[title].active.title = title 
            dict_wbtitle[title][title].append(fi_row)

#save
    wb0.save('courses.xlsx')
    for k in dict_wbtitle.keys():
        dict_wbtitle[k].save(k+'.xlsx')
if __name__ == '__main__':
    rr()
